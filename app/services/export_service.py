"""Exportação de dados para Excel (.xlsx).

Gera um workbook com múltiplas abas a partir dos dados consolidados do
período selecionado. O objetivo é substituir a planilha de controle que
a barbearia usava antes — com os mesmos números que aparecem no
Dashboard, mas "destacáveis" para mandar por email, imprimir ou
analisar com pivot no Excel.

Abas geradas (nesta ordem):
  1. Resumo            — KPIs do período e do mês corrente
  2. Receitas          — lista detalhada com servico, valor, forma, cliente
  3. Despesas          — lista detalhada com categoria, descrição, valor
  4. Serie diaria      — receita x despesa x saldo por dia
  5. Ranking servicos  — serviços ordenados por faturamento
  6. Mix pagamento     — forma de pagamento (receita)
  7. Despesas x cat.   — agregado de despesas por categoria

Valores em centavos são convertidos para reais (float) para que o
usuário possa aplicar SOMA / médias no Excel sem ficar dividindo por
100. A formatação de célula é "Contábil pt-BR" (R$).
"""

from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services import configuracao_service, despesa_service, receita_service
from app.utils.formato import centavos_para_reais
from app.utils.validators import ValidacaoError

# ---------------------------------------------------------------------------
# Constantes de estilo — mantidas centralizadas para consistência visual.
# ---------------------------------------------------------------------------

_FONTE = "Arial"
_TITULO_TAM = 16
_HEADER_TAM = 11
_CORPO_TAM = 10

_COR_TITULO_FG = "FFFFFF"
_COR_TITULO_BG = "1F2937"       # cinza-escuro
_COR_HEADER_FG = "FFFFFF"
_COR_HEADER_BG = "D4A14A"       # dourado Brodz
_COR_TOTAL_BG = "F3F4F6"        # cinza claro
_COR_ZEBRA = "FAFAFA"

_FMT_MOEDA = 'R$ #,##0.00;[Red]-R$ #,##0.00;"-"'
_FMT_INT = '#,##0'
_FMT_PCT = '0.0%'
_FMT_DATA = "dd/mm/yyyy"

_BORDA_FINA = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


# ---------------------------------------------------------------------------
# Função pública
# ---------------------------------------------------------------------------


def exportar_periodo_xlsx(
    conn: sqlite3.Connection,
    de: date,
    ate: date,
    caminho_saida: Path | str,
) -> Path:
    """Gera o workbook XLSX para o período e salva em ``caminho_saida``.

    Retorna o Path efetivo do arquivo gerado.
    Levanta ``ValidacaoError`` se ``de > ate``.
    """
    if de > ate:
        raise ValidacaoError("A data inicial deve ser anterior ou igual à final.")

    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    # Carrega todos os dados de uma vez — mais rápido que abrir cursor múltiplas vezes.
    receitas = receita_service.listar_periodo(conn, de, ate)
    despesas = despesa_service.listar_periodo(conn, de, ate)
    serie_r = receita_service.serie_diaria_receita(conn, de, ate)
    serie_d = despesa_service.serie_diaria_despesa(conn, de, ate)
    ranking = receita_service.ranking_servicos(conn, de, ate, limite=50)
    mix = receita_service.mix_forma_pagamento(conn, de, ate)
    por_cat = despesa_service.total_por_categoria(conn, de, ate)

    totais_r = receita_service.totais_periodo(conn, de, ate)
    totais_d = despesa_service.totais_periodo(conn, de, ate)

    # Meta sempre olha o mês corrente (mesma regra do Dashboard).
    hoje = date.today()
    primeiro_mes = hoje.replace(day=1)
    meta_centavos = configuracao_service.meta_mensal_centavos(conn)
    totais_mes = receita_service.totais_periodo(conn, primeiro_mes, hoje)

    wb = Workbook()
    # Remove a aba default para que a primeira criada seja "Resumo".
    default = wb.active
    wb.remove(default)

    _aba_resumo(
        wb, de, ate,
        totais_r=totais_r, totais_d=totais_d,
        meta_centavos=meta_centavos,
        receita_mes_centavos=totais_mes["total_centavos"],
    )
    _aba_receitas(wb, receitas)
    _aba_despesas(wb, despesas)
    _aba_serie_diaria(wb, serie_r, serie_d)
    _aba_ranking(wb, ranking)
    _aba_mix(wb, mix)
    _aba_por_categoria(wb, por_cat)

    wb.save(str(caminho_saida))
    return caminho_saida


# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------


def _aplicar_titulo(ws: Worksheet, texto: str, col_final: int = 6) -> None:
    """Linha 1: título grande da aba, mesclado do A1 até ``col_final``."""
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_final)
    cel = ws.cell(row=1, column=1, value=texto)
    cel.font = Font(name=_FONTE, size=_TITULO_TAM, bold=True, color=_COR_TITULO_FG)
    cel.fill = PatternFill("solid", fgColor=_COR_TITULO_BG)
    cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _escrever_header(
    ws: Worksheet, linha: int, colunas: Sequence[str], largura_padrao: int = 18
) -> None:
    fill = PatternFill("solid", fgColor=_COR_HEADER_BG)
    font = Font(name=_FONTE, size=_HEADER_TAM, bold=True, color=_COR_HEADER_FG)
    align = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = 22
    for idx, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=linha, column=idx, value=nome)
        cel.font = font
        cel.fill = fill
        cel.alignment = align
        cel.border = _BORDA_FINA
        # Largura: heurística — usa o maior entre padrão e tamanho do texto + 2.
        ws.column_dimensions[get_column_letter(idx)].width = max(
            largura_padrao, len(nome) + 2
        )


def _escrever_linhas(
    ws: Worksheet,
    inicio_linha: int,
    linhas: Iterable[Sequence],
    *,
    formatos: Sequence[str | None] | None = None,
    zebra: bool = True,
) -> int:
    """Escreve ``linhas`` a partir de ``inicio_linha``. Retorna a próxima linha livre."""
    font = Font(name=_FONTE, size=_CORPO_TAM)
    fill_zebra = PatternFill("solid", fgColor=_COR_ZEBRA)
    lin = inicio_linha
    for i, linha in enumerate(linhas):
        for col_idx, valor in enumerate(linha, start=1):
            cel = ws.cell(row=lin, column=col_idx, value=valor)
            cel.font = font
            cel.border = _BORDA_FINA
            if formatos and col_idx - 1 < len(formatos) and formatos[col_idx - 1]:
                cel.number_format = formatos[col_idx - 1]  # type: ignore[assignment]
            if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                cel.alignment = Alignment(horizontal="right")
            elif isinstance(valor, (date, datetime)):
                cel.alignment = Alignment(horizontal="center")
            else:
                cel.alignment = Alignment(horizontal="left", wrap_text=False)
        if zebra and i % 2 == 1:
            for col_idx in range(1, (len(linha) or 1) + 1):
                ws.cell(row=lin, column=col_idx).fill = fill_zebra
        lin += 1
    return lin


def _escrever_totais(
    ws: Worksheet,
    linha: int,
    rotulo: str,
    rotulo_col: int,
    formulas: dict[int, str],
) -> None:
    """Linha de totais destacada. ``formulas`` mapeia coluna -> fórmula Excel."""
    font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
    fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
    ws.cell(row=linha, column=rotulo_col, value=rotulo).font = font
    for col in range(1, max(formulas.keys(), default=rotulo_col) + 1):
        ws.cell(row=linha, column=col).fill = fill
        ws.cell(row=linha, column=col).border = _BORDA_FINA
        ws.cell(row=linha, column=col).font = font
    for col, formula in formulas.items():
        cel = ws.cell(row=linha, column=col, value=formula)
        cel.number_format = _FMT_MOEDA
        cel.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------


def _aba_resumo(
    wb: Workbook,
    de: date,
    ate: date,
    *,
    totais_r: dict,
    totais_d: dict,
    meta_centavos: int,
    receita_mes_centavos: int,
) -> None:
    ws = wb.create_sheet("Resumo")
    _aplicar_titulo(ws, f"Fluxo Barber — Resumo ({de.strftime('%d/%m/%Y')} a {ate.strftime('%d/%m/%Y')})", col_final=4)

    receita_reais = centavos_para_reais(totais_r["total_centavos"])
    despesa_reais = centavos_para_reais(totais_d["total_centavos"])
    qtd_atend = int(totais_r["qtd_atendimentos"])
    # Linhas (indicador, valor, formato, observação)
    dados: list[tuple[str, float | int | str, str | None, str]] = [
        ("Receita total",   receita_reais,                 _FMT_MOEDA, "soma dos atendimentos no período"),
        ("Despesas totais", despesa_reais,                 _FMT_MOEDA, "soma dos lançamentos de despesa"),
        ("Saldo",           "=B3-B4",                      _FMT_MOEDA, "receita menos despesa"),
        ("Atendimentos",    qtd_atend,                     _FMT_INT,   "quantidade de receitas registradas"),
        ("Ticket médio",    "=IFERROR(B3/B6,0)",           _FMT_MOEDA, "receita / atendimentos"),
        ("",                "",                            None,       ""),
        ("Meta mensal",     centavos_para_reais(meta_centavos), _FMT_MOEDA, "configurada no app"),
        ("Receita no mês",  centavos_para_reais(receita_mes_centavos), _FMT_MOEDA, f"mês corrente — {date.today().strftime('%m/%Y')}"),
        ("% da meta",       "=IFERROR(B10/B9,0)",          _FMT_PCT,   "progresso em relação à meta"),
    ]

    # Header
    _escrever_header(ws, 2, ["Indicador", "Valor", "", "Observação"], largura_padrao=22)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 50

    font_corpo = Font(name=_FONTE, size=_CORPO_TAM)
    font_bold = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
    for i, (rotulo, valor, fmt, obs) in enumerate(dados):
        lin = 3 + i
        ws.cell(row=lin, column=1, value=rotulo).font = font_bold
        cel = ws.cell(row=lin, column=2, value=valor)
        cel.font = font_corpo
        cel.alignment = Alignment(horizontal="right")
        if fmt:
            cel.number_format = fmt
        ws.cell(row=lin, column=4, value=obs).font = Font(
            name=_FONTE, size=_CORPO_TAM, italic=True, color="6B7280"
        )
        ws.cell(row=lin, column=1).border = _BORDA_FINA
        ws.cell(row=lin, column=2).border = _BORDA_FINA

    ws.sheet_view.showGridLines = False


def _aba_receitas(wb: Workbook, receitas: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Receitas")
    _aplicar_titulo(ws, "Receitas — detalhe", col_final=7)
    _escrever_header(
        ws, 2,
        ["Data", "Serviço", "Cliente", "Plano", "Forma pgto.", "Valor (R$)", "Observação"],
        largura_padrao=18,
    )

    linhas = []
    for r in receitas:
        linhas.append([
            date.fromisoformat(r["data"]) if r["data"] else None,
            r["servico_nome"] or "",
            r["cliente_nome"] or "",
            r["plano_nome"] or "",
            r["forma_pagamento"] or "—",
            centavos_para_reais(int(r["valor"] or 0)),
            r["observacao"] or "",
        ])

    formatos = [_FMT_DATA, None, None, None, None, _FMT_MOEDA, None]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        primeira = 3
        ultima = proximo - 1
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=1,
            formulas={6: f"=SUM(F{primeira}:F{ultima})"},
        )
        # Rótulo de contagem na coluna A+1
        cel_qtd = ws.cell(row=proximo, column=2, value=f"{len(linhas)} atendimento(s)")
        cel_qtd.font = Font(name=_FONTE, size=_CORPO_TAM, italic=True, color="6B7280")
        cel_qtd.alignment = Alignment(horizontal="left")

    # Autofilter + freeze
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{max(2, 2 + len(linhas))}"
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["G"].width = 40


def _aba_despesas(wb: Workbook, despesas: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Despesas")
    _aplicar_titulo(ws, "Despesas — detalhe", col_final=5)
    _escrever_header(
        ws, 2,
        ["Data", "Categoria", "Descrição", "Forma pgto.", "Valor (R$)"],
        largura_padrao=18,
    )

    linhas = []
    for d in despesas:
        linhas.append([
            date.fromisoformat(d["data"]) if d["data"] else None,
            d["categoria_nome"] or "",
            d["descricao"] or "",
            d["forma_pagamento"] or "",
            centavos_para_reais(int(d["valor"] or 0)),
        ])

    formatos = [_FMT_DATA, None, None, None, _FMT_MOEDA]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=1,
            formulas={5: f"=SUM(E3:E{proximo - 1})"},
        )
        ws.cell(row=proximo, column=2, value=f"{len(linhas)} despesa(s)").font = Font(
            name=_FONTE, size=_CORPO_TAM, italic=True, color="6B7280"
        )

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{max(2, 2 + len(linhas))}"
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 36


def _aba_serie_diaria(
    wb: Workbook,
    serie_r: list[dict],
    serie_d: list[dict],
) -> None:
    ws = wb.create_sheet("Serie diaria")
    _aplicar_titulo(ws, "Evolução diária — receita x despesa", col_final=5)
    _escrever_header(
        ws, 2,
        ["Data", "Receita (R$)", "Despesa (R$)", "Saldo (R$)", "Atendimentos"],
        largura_padrao=18,
    )

    # Une as duas séries por data — datas são strings ISO.
    por_data_r = {r["data"]: r for r in serie_r}
    por_data_d = {r["data"]: r for r in serie_d}
    todas = sorted(set(por_data_r.keys()) | set(por_data_d.keys()))

    linhas = []
    for i, data_iso in enumerate(todas):
        r = por_data_r.get(data_iso)
        d = por_data_d.get(data_iso)
        # Saldo é formulado para que o usuário possa editar valores e recalcular.
        linha_excel = 3 + i
        linhas.append([
            date.fromisoformat(data_iso),
            centavos_para_reais(int(r["total"])) if r else 0,
            centavos_para_reais(int(d["total"])) if d else 0,
            f"=B{linha_excel}-C{linha_excel}",
            int(r["qtd"]) if r else 0,
        ])

    formatos = [_FMT_DATA, _FMT_MOEDA, _FMT_MOEDA, _FMT_MOEDA, _FMT_INT]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        ultima = proximo - 1
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=1,
            formulas={
                2: f"=SUM(B3:B{ultima})",
                3: f"=SUM(C3:C{ultima})",
                4: f"=SUM(D3:D{ultima})",
            },
        )
        # Atendimentos total como inteiro
        cel_q = ws.cell(row=proximo, column=5, value=f"=SUM(E3:E{ultima})")
        cel_q.number_format = _FMT_INT
        cel_q.font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
        cel_q.alignment = Alignment(horizontal="right")
        cel_q.fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
        cel_q.border = _BORDA_FINA

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{max(2, 2 + len(linhas))}"


def _aba_ranking(wb: Workbook, ranking: list[dict]) -> None:
    ws = wb.create_sheet("Ranking servicos")
    _aplicar_titulo(ws, "Ranking de serviços por faturamento", col_final=5)
    _escrever_header(
        ws, 2,
        ["#", "Serviço", "Atendimentos", "Total (R$)", "Ticket médio (R$)"],
        largura_padrao=18,
    )

    linhas = []
    for pos, r in enumerate(ranking, start=1):
        qtd = int(r["qtd"] or 0)
        total = int(r["total"] or 0)
        linha_excel = 2 + pos  # posição no Excel
        linhas.append([
            pos,
            r["servico_nome"] or "",
            qtd,
            centavos_para_reais(total),
            f"=IFERROR(D{linha_excel}/C{linha_excel},0)",
        ])

    formatos = [_FMT_INT, None, _FMT_INT, _FMT_MOEDA, _FMT_MOEDA]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        ultima = proximo - 1
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=2,
            formulas={4: f"=SUM(D3:D{ultima})"},
        )
        cel_qtd = ws.cell(row=proximo, column=3, value=f"=SUM(C3:C{ultima})")
        cel_qtd.number_format = _FMT_INT
        cel_qtd.font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
        cel_qtd.alignment = Alignment(horizontal="right")
        cel_qtd.fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
        cel_qtd.border = _BORDA_FINA

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.freeze_panes = "A3"


def _aba_mix(wb: Workbook, mix: list[dict]) -> None:
    ws = wb.create_sheet("Mix pagamento")
    _aplicar_titulo(ws, "Receita por forma de pagamento", col_final=4)
    _escrever_header(
        ws, 2,
        ["Forma", "Atendimentos", "Total (R$)", "% do período"],
        largura_padrao=18,
    )

    total_geral = sum(int(m["total"] or 0) for m in mix)

    linhas = []
    for i, m in enumerate(mix):
        qtd = int(m["qtd"] or 0)
        total = int(m["total"] or 0)
        linha_excel = 3 + i
        linhas.append([
            m["forma"] or "—",
            qtd,
            centavos_para_reais(total),
            f"=IFERROR(C{linha_excel}/$C${3 + len(mix)},0)" if mix else 0,
        ])

    formatos = [None, _FMT_INT, _FMT_MOEDA, _FMT_PCT]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        ultima = proximo - 1
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=1,
            formulas={3: f"=SUM(C3:C{ultima})"},
        )
        cel_pct = ws.cell(row=proximo, column=4, value="=IFERROR(C{row}/C{row},0)".format(row=proximo))
        cel_pct.number_format = _FMT_PCT
        cel_pct.font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
        cel_pct.alignment = Alignment(horizontal="right")
        cel_pct.fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
        cel_pct.border = _BORDA_FINA

        cel_qtd = ws.cell(row=proximo, column=2, value=f"=SUM(B3:B{ultima})")
        cel_qtd.number_format = _FMT_INT
        cel_qtd.font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
        cel_qtd.alignment = Alignment(horizontal="right")
        cel_qtd.fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
        cel_qtd.border = _BORDA_FINA

    ws.freeze_panes = "A3"


def _aba_por_categoria(wb: Workbook, por_cat: list[dict]) -> None:
    ws = wb.create_sheet("Despesas por cat.")
    _aplicar_titulo(ws, "Despesas por categoria", col_final=4)
    _escrever_header(
        ws, 2,
        ["Categoria", "Lançamentos", "Total (R$)", "% do período"],
        largura_padrao=22,
    )

    linhas = []
    total_rows = len(por_cat)
    for i, c in enumerate(por_cat):
        qtd = int(c["qtd"] or 0)
        total = int(c["total"] or 0)
        linha_excel = 3 + i
        pct = (
            f"=IFERROR(C{linha_excel}/$C${3 + total_rows},0)"
            if total_rows else 0
        )
        linhas.append([
            c["categoria_nome"] or "",
            qtd,
            centavos_para_reais(total),
            pct,
        ])

    formatos = [None, _FMT_INT, _FMT_MOEDA, _FMT_PCT]
    proximo = _escrever_linhas(ws, 3, linhas, formatos=formatos)

    if linhas:
        ultima = proximo - 1
        _escrever_totais(
            ws, proximo,
            rotulo="TOTAL", rotulo_col=1,
            formulas={3: f"=SUM(C3:C{ultima})"},
        )
        cel_qtd = ws.cell(row=proximo, column=2, value=f"=SUM(B3:B{ultima})")
        cel_qtd.number_format = _FMT_INT
        cel_qtd.font = Font(name=_FONTE, size=_CORPO_TAM, bold=True)
        cel_qtd.alignment = Alignment(horizontal="right")
        cel_qtd.fill = PatternFill("solid", fgColor=_COR_TOTAL_BG)
        cel_qtd.border = _BORDA_FINA

    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 30


