"""Testes do export XLSX.

Valida que o arquivo gerado:
  - contém as 7 abas esperadas, na ordem correta;
  - tem os KPIs numericamente coerentes com os serviços agregados;
  - persiste as fórmulas (não hardcoda soma/saldo/%);
  - lida com período vazio (ainda gera arquivo com headers);
  - rejeita de > ate.
"""

from __future__ import annotations

import sqlite3
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.enums import FormaPagamento
from app.services import despesa_service, export_service, receita_service
from app.utils.validators import ValidacaoError


ABAS_ESPERADAS = [
    "Resumo",
    "Receitas",
    "Despesas",
    "Serie diaria",
    "Ranking servicos",
    "Mix pagamento",
    "Despesas por cat.",
]


def _sid(conn: sqlite3.Connection, nome: str) -> int:
    return conn.execute("SELECT id FROM servico WHERE nome = ?", (nome,)).fetchone()["id"]


def _cid(conn: sqlite3.Connection, nome: str) -> int:
    return conn.execute(
        "SELECT id FROM categoria_despesa WHERE nome = ?", (nome,)
    ).fetchone()["id"]


# ---------------------------------------------------------------------------


def test_export_periodo_gera_arquivo_com_todas_as_abas(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    hoje = date.today()
    sid = _sid(conn, "Corte de cabelo")
    cid = _cid(conn, "Outros")
    receita_service.registrar_atendimento_avulso(
        conn, servico_id=sid, valor_centavos=4000,
        forma_pagamento=FormaPagamento.PIX,
    )
    despesa_service.registrar_despesa(
        conn, categoria_id=cid, descricao="Teste",
        valor_centavos=1500, forma_pagamento=FormaPagamento.PIX,
    )

    saida = tmp_path / "out.xlsx"
    resultado = export_service.exportar_periodo_xlsx(conn, hoje, hoje, saida)

    assert resultado == saida
    assert saida.exists()
    assert saida.stat().st_size > 0

    wb = load_workbook(saida)
    assert wb.sheetnames == ABAS_ESPERADAS


def test_export_resumo_tem_kpis_coerentes(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    hoje = date.today()
    sid = _sid(conn, "Corte de cabelo")
    cid = _cid(conn, "Outros")
    # R$ 100 de receita (2 atend) e R$ 30 de despesa
    for _ in range(2):
        receita_service.registrar_atendimento_avulso(
            conn, servico_id=sid, valor_centavos=5000,
            forma_pagamento=FormaPagamento.PIX,
        )
    despesa_service.registrar_despesa(
        conn, categoria_id=cid, descricao="x",
        valor_centavos=3000, forma_pagamento=FormaPagamento.PIX,
    )

    saida = tmp_path / "resumo.xlsx"
    export_service.exportar_periodo_xlsx(conn, hoje, hoje, saida)

    wb = load_workbook(saida)
    ws = wb["Resumo"]
    # Linhas 3..11 são os indicadores (ver export_service._aba_resumo).
    # Receita total e Despesas totais são valores, não fórmulas.
    assert ws["A3"].value == "Receita total"
    assert ws["B3"].value == 100.0  # R$ 100,00
    assert ws["A4"].value == "Despesas totais"
    assert ws["B4"].value == 30.0
    # Saldo é fórmula
    assert ws["A5"].value == "Saldo"
    assert ws["B5"].value == "=B3-B4"
    # Atendimentos inteiro
    assert ws["A6"].value == "Atendimentos"
    assert ws["B6"].value == 2
    # Ticket médio é fórmula
    assert ws["B7"].value == "=IFERROR(B3/B6,0)"


def test_export_receitas_inclui_detalhe_e_formula_de_total(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    hoje = date.today()
    sid = _sid(conn, "Barba")
    for valor in (2500, 3500, 4500):
        receita_service.registrar_atendimento_avulso(
            conn, servico_id=sid, valor_centavos=valor,
            forma_pagamento=FormaPagamento.DINHEIRO,
        )

    saida = tmp_path / "rec.xlsx"
    export_service.exportar_periodo_xlsx(conn, hoje, hoje, saida)
    wb = load_workbook(saida)
    ws = wb["Receitas"]

    # Linha 1 título, linha 2 header, 3..5 dados, 6 total.
    assert ws["A1"].value.startswith("Receitas")
    assert ws["A2"].value == "Data"
    assert ws["F3"].value + ws["F4"].value + ws["F5"].value == pytest.approx(25 + 35 + 45)
    # A linha de total tem fórmula SUM
    assert ws["A6"].value == "TOTAL"
    assert ws["F6"].value == "=SUM(F3:F5)"


def test_export_serie_diaria_usa_formula_para_saldo(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    hoje = date.today()
    ontem = hoje - timedelta(days=1)
    sid = _sid(conn, "Corte de cabelo")
    cid = _cid(conn, "Outros")
    receita_service.registrar_atendimento_avulso(
        conn, servico_id=sid, valor_centavos=5000,
        forma_pagamento=FormaPagamento.PIX, data_atendimento=ontem,
    )
    despesa_service.registrar_despesa(
        conn, categoria_id=cid, descricao="x", valor_centavos=2000,
        forma_pagamento=FormaPagamento.PIX, data_despesa=ontem,
    )
    receita_service.registrar_atendimento_avulso(
        conn, servico_id=sid, valor_centavos=7000,
        forma_pagamento=FormaPagamento.PIX,
    )

    saida = tmp_path / "serie.xlsx"
    export_service.exportar_periodo_xlsx(conn, ontem, hoje, saida)
    wb = load_workbook(saida)
    ws = wb["Serie diaria"]

    # Linha 3 é o primeiro dia (ontem). Saldo tem que ser fórmula B3-C3.
    assert ws["D3"].value == "=B3-C3"
    assert ws["D4"].value == "=B4-C4"
    # Valores numéricos de receita correspondem aos lançamentos.
    assert ws["B3"].value == 50.0
    assert ws["C3"].value == 20.0
    assert ws["B4"].value == 70.0


def test_export_rejeita_periodo_invertido(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    with pytest.raises(ValidacaoError, match="anterior"):
        export_service.exportar_periodo_xlsx(
            conn, date(2026, 5, 1), date(2026, 4, 1),
            tmp_path / "invalido.xlsx",
        )


def test_export_periodo_vazio_ainda_gera_arquivo_valido(
    conn: sqlite3.Connection, tmp_path: Path
) -> None:
    """Se não há receitas/despesas no período, o arquivo deve existir
    com todas as abas, mesmo sem linhas de dados."""
    saida = tmp_path / "vazio.xlsx"
    export_service.exportar_periodo_xlsx(conn, date(2020, 1, 1), date(2020, 1, 1), saida)
    wb = load_workbook(saida)
    assert wb.sheetnames == ABAS_ESPERADAS
    # Receitas vazia tem só título + header
    ws_rec = wb["Receitas"]
    assert ws_rec.cell(row=3, column=1).value is None
    # Resumo mesmo vazio tem os labels
    ws_res = wb["Resumo"]
    assert ws_res["A3"].value == "Receita total"
    assert ws_res["B3"].value == 0
